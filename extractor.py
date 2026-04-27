"""
Enhanced Ad-hoc Extractor v2
════════════════════════════
Pulls candidate profiles from a mailbox (from 2026-01-01), searches the
full reply chain of every thread, parses HTML tables with fuzzy/alias
header matching, and falls back to a free AI (Google Gemini or Ollama)
when:
  • No HTML table is found, OR
  • A table is found but the headers cannot be mapped to known fields

Output: adhoc_candidates_<timestamp>.xlsx  +  PostgreSQL insert

Subject filters accepted:
  - CODE: ...          e.g.  BS: Salesforce Developer | 31509
  - [External]: CODE: ...

Attachments (PDF/DOC/DOCX) are uploaded to OneDrive Archive folder.

Required env vars:
  AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET
  TARGET_MAILBOX
  OD_TENANT_ID, OD_CLIENT_ID, OD_REFRESH_TOKEN, ONEDRIVE_USER
  DB_DSN

Optional:
  OD_CLIENT_SECRET            (OneDrive secret — can be empty for public apps)
  INBOX_SUBFOLDER             (default: "Company Profiles")
  DB_SCHEMA                   (default: "public")
  DB_TABLE_NAME               (default: "temp_hrvolibit_archive")
  ONEDRIVE_ARCHIVE_FOLDER     (default: "archive")
  LIMIT                       (default: 100)
  SKIP_UPLOADS                (default: false)
  AI_PROVIDER                 (default: "gemini" — options: gemini | ollama | none)
  GEMINI_API_KEY              (required when AI_PROVIDER=gemini)
  OLLAMA_URL                  (default: http://localhost:11434 for local Ollama)
  OLLAMA_MODEL                (default: llama3)
  ONLY_COMPLETE               (default: false — set "true" to skip rows missing key fields)

── Local run ─────────────────────────────────────────────────────────────────
1. Install:
       pip install msal openpyxl requests python-dotenv psycopg2-binary \
                   google-generativeai rapidfuzz

2. Create .env (never commit):
       AZURE_TENANT_ID=...
       AZURE_CLIENT_ID=...
       AZURE_CLIENT_SECRET=...
       TARGET_MAILBOX=mailbox@yourdomain.com
       OD_TENANT_ID=...
       OD_CLIENT_ID=...
       OD_REFRESH_TOKEN=...
       ONEDRIVE_USER=user@yourdomain.com
       DB_DSN=postgresql://user:password@host:5432/dbname
       AI_PROVIDER=gemini
       GEMINI_API_KEY=your-free-gemini-api-key

3. Run:
       python bs_extractor_v2.py
──────────────────────────────────────────────────────────────────────────────
"""

import os
import re
import html as html_lib
import logging
import json
import requests
import psycopg2
from psycopg2 import sql as pgsql
from datetime import datetime, date
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from msal import ConfidentialClientApplication, PublicClientApplication
from dotenv import load_dotenv

load_dotenv()

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler("adhoc_extract_v2.log")],
)
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
AZURE_TENANT_ID     = os.environ["AZURE_TENANT_ID"]
AZURE_CLIENT_ID     = os.environ["AZURE_CLIENT_ID"]
AZURE_CLIENT_SECRET = os.environ["AZURE_CLIENT_SECRET"]

TARGET_MAILBOX  = os.environ["TARGET_MAILBOX"]
INBOX_SUBFOLDER = os.environ.get("INBOX_SUBFOLDER", "Company Profiles")
VOLIBITS_DOMAIN = "volibits.com"
RESUME_EXTENSIONS = {".pdf", ".doc", ".docx"}

OD_TENANT_ID     = os.environ["OD_TENANT_ID"]
OD_CLIENT_ID     = os.environ["OD_CLIENT_ID"]
OD_CLIENT_SECRET = os.environ.get("OD_CLIENT_SECRET", "")
OD_REFRESH_TOKEN = os.environ["OD_REFRESH_TOKEN"]
ONEDRIVE_USER    = os.environ["ONEDRIVE_USER"]
DB_DSN           = os.environ["DB_DSN"]

DB_SCHEMA = os.environ.get("DB_SCHEMA", "public")
DB_TABLE  = os.environ.get("DB_TABLE_NAME", "temp_hrvolibit_archive")

ONEDRIVE_ARCHIVE_FOLDER = os.environ.get("ONEDRIVE_ARCHIVE_FOLDER", "archive")
LIMIT        = int(os.environ.get("LIMIT", "100"))
SKIP_UPLOADS = os.environ.get("SKIP_UPLOADS", "false").strip().lower() in ("1", "true", "yes")
ONLY_COMPLETE = os.environ.get("ONLY_COMPLETE", "false").strip().lower() in ("1", "true", "yes")

# ── AI Config ──────────────────────────────────────────────────────────────────
AI_PROVIDER    = os.environ.get("AI_PROVIDER", "gemini").strip().lower()
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "AIzaSyB0AR_FPfmCU0vxSZCJxce_O26q5ZL5TIE")
OLLAMA_URL     = os.environ.get("OLLAMA_URL", "http://localhost:11434")
OLLAMA_MODEL   = os.environ.get("OLLAMA_MODEL", "llama3")

SINCE_DATE = date(2026, 1, 1)

# ── Company codes ──────────────────────────────────────────────────────────────
COMPANY_CODES: dict[str, str] = {
    "BS":  "Birlasoft",
    "BW":  "BeWealthy",
    "TCS": "TCS",
    "INF": "Infosys",
    "WIP": "Wipro",
    "HCL": "HCL Technologies",
    "ACC": "Accenture",
    "CAP": "Capgemini",
    "COG": "Cognizant",
    "IBM": "IBM",
    "SAP": "SAP",
    "ORC": "Oracle",
    "MS":  "Microsoft",
    "RS":  "RS Software",
}

# ── Columns ────────────────────────────────────────────────────────────────────
EXCEL_COLUMNS = [
    "email_subject",
    "recruiter", "client_recruiter", "email_from", "email_to",
    "delivery_type", "company_name", "general_skill",
    "date", "jr_no",
    "name_of_candidate", "contact_number", "email_id",
    "total_experience", "relevant_experience",
    "current_ctc", "expected_ctc", "notice_period",
    "current_org", "current_location", "preferred_location",
    "qualification", "attachment", "remarks", "record_status",
    "ai_extracted",
]

DB_COLUMNS = [c for c in EXCEL_COLUMNS if c not in ("ai_extracted",)]

# ── Column aliases (expanded to match the image headers) ──────────────────────
COLUMN_ALIASES: dict[str, list[str]] = {
    "general_skill": [
        "general_skill", "generalskill", "gen_skill", "genskill",
        "skill", "requirement", "skil", "skills", "technology",
        "tech", "requirement skill",
    ],
    "jr_no": [
        "jr no", "jr_no", "jr no.", "jr", "jrno", "req_id", "req id",
        "requisition id", "requisition_id", "jr_number", "jr number",
        "job req", "job requisition", "job_req", "reqid",
        "requirement id", "requirement_id", "jrno.", "jr#",
        "jr no(mention the jr number where profiles are uploaded on sf).",
    ],
    "name_of_candidate": [
        "candidate name", "candidate_name", "name", "applicant name",
        "applicant_name", "full name", "full_name", "name of candidate",
        "candidate", "person name", "n_of_candidate", "cand name",
        "candidate name", "first name", "candidatename",
    ],
    "contact_number": [
        "contact number", "contact_number", "phone", "mobile", "cell",
        "phone number", "phone_number", "mobile number", "mobile_number",
        "contact no", "contact_no", "ph no", "ph_no", "contact",
        "mobile no.", "contact details", "contact detail",
        "mob no", "mob", "phone no",
    ],
    "email_id": [
        "email id", "email_id", "email", "e-mail", "mail id", "mail_id",
        "email address", "email_address", "e mail", "e_id", "emailid",
        "e-mail id", "email address", "mail",
    ],
    "current_org": [
        "current company", "current_company", "company", "employer",
        "current employer", "current_employer", "current org", "current_org",
        "organisation", "organization", "curr company", "present company",
        "current / last company", "organization", "org",
    ],
    "total_experience": [
        "total experience", "total_experience", "total exp", "total_exp",
        "experience", "exp", "total yrs", "total years", "tot exp",
        "total exp (yrs)", "total experience (yrs)", "tot_exp",
        "total exp.", "yrs of exp",
    ],
    "relevant_experience": [
        "relevant experience", "relevant_experience", "relevant exp",
        "relevant_exp", "rel exp", "rel_exp", "relevant yrs",
        "rel exp.", "rel experience", "rel. exp", "rel.exp",
        "relevant exp (yrs)", "rel exp (yrs)",
    ],
    "current_ctc": [
        "current ctc", "current_ctc", "ctc", "current salary",
        "current_salary", "present ctc", "present_ctc", "curr ctc",
        "cctc", "currctc", "curr_ctc", "c_ctc",
        "current ctc (lakhs)", "ctcc", "cur ctc",
    ],
    "expected_ctc": [
        "expected ctc", "expected_ctc", "expected salary", "expected_salary",
        "exp ctc", "exp_ctc", "desired ctc", "desired_ctc",
        "ectc", "e_ctc", "expctc", "ect",
        "exp ctc (lakhs)", "expected ctc (lakhs)", "exp. ctc",
    ],
    "notice_period": [
        "notice period", "notice_period", "notice", "np", "joining time",
        "notice days", "notice_days", "notice period(days)", "notice (days)",
        "notice period (days)", "nperiod",
    ],
    "current_location": [
        "current location", "current_location", "location", "city",
        "current city", "current_city", "present location",
        "current_loc", "curr location", "c loc", "c_location",
        "curr_location", "curr_loc", "current/preferred location",
    ],
    "preferred_location": [
        "preferred location", "preferred_location", "preferred city",
        "preferred_city", "pref location", "pref_location",
        "willing to relocate", "target location", "pref_loc",
        "pre location", "current/preferred location", "pref location",
    ],
    "qualification": [
        "qualification", "qualifications", "education", "degree",
        "educational qualification", "highest qualification",
        "academic qualification", "edu", "qual",
    ],
    "date": [
        "date", "submission date", "submission_date", "applied date",
        "applied_date", "profile date", "date of submission",
    ],
    "remarks": [
        "remarks", "comments", "comment", "note", "notes",
        "remarks/availability for interview", "remark",
        "availability", "availability for interview",
    ],
}

def _normalize(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())

ALIAS_MAP: dict[str, str] = {
    _normalize(alias): canonical
    for canonical, aliases in COLUMN_ALIASES.items()
    for alias in aliases
}

def _resolve_header(raw: str) -> Optional[str]:
    key = _normalize(raw)
    if key in ALIAS_MAP:
        return ALIAS_MAP[key]
    # Fuzzy match using rapidfuzz if installed
    try:
        from rapidfuzz import process, fuzz
        best = process.extractOne(key, ALIAS_MAP.keys(), scorer=fuzz.token_sort_ratio)
        if best and best[1] >= 75:
            return ALIAS_MAP[best[0]]
    except ImportError:
        pass
    return None


# ── AI Extraction ──────────────────────────────────────────────────────────────
AI_EXTRACTION_PROMPT = """You are a data extraction assistant. Extract candidate profile information from the following email body text.

Return ONLY a valid JSON array (no markdown, no explanation). Each element is a candidate object with these keys (use null if not found):
- name_of_candidate
- contact_number
- email_id
- total_experience
- relevant_experience
- current_ctc
- expected_ctc
- notice_period
- current_org
- current_location
- preferred_location
- qualification
- remarks

Email body:
{body}
"""

def _ai_extract_gemini(body_text: str) -> list[dict]:
    """Use Google Gemini free API to extract candidates from plain text."""
    if not GEMINI_API_KEY:
        log.warning("GEMINI_API_KEY not set — skipping Gemini extraction.")
        return []
    prompt = AI_EXTRACTION_PROMPT.format(body=body_text[:8000])
    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"gemini-1.5-flash:generateContent?key={GEMINI_API_KEY}"
    )
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0, "maxOutputTokens": 2048},
    }
    try:
        resp = requests.post(url, json=payload, timeout=30)
        resp.raise_for_status()
        raw = resp.json()
        text = raw["candidates"][0]["content"]["parts"][0]["text"]
        # Strip markdown code fences if present
        text = re.sub(r"```(?:json)?", "", text).strip().strip("`").strip()
        data = json.loads(text)
        if isinstance(data, dict):
            data = [data]
        log.info(f"  [Gemini] extracted {len(data)} candidate(s) via AI.")
        return data if isinstance(data, list) else []
    except Exception as exc:
        log.warning(f"Gemini extraction failed: {exc}")
        return []


def _ai_extract_ollama(body_text: str) -> list[dict]:
    """Use local Ollama (free) to extract candidates from plain text."""
    prompt = AI_EXTRACTION_PROMPT.format(body=body_text[:6000])
    try:
        resp = requests.post(
            f"{OLLAMA_URL}/api/generate",
            json={"model": OLLAMA_MODEL, "prompt": prompt, "stream": False},
            timeout=120,
        )
        resp.raise_for_status()
        text = resp.json().get("response", "")
        text = re.sub(r"```(?:json)?", "", text).strip().strip("`").strip()
        data = json.loads(text)
        if isinstance(data, dict):
            data = [data]
        log.info(f"  [Ollama] extracted {len(data)} candidate(s) via AI.")
        return data if isinstance(data, list) else []
    except Exception as exc:
        log.warning(f"Ollama extraction failed: {exc}")
        return []


def ai_extract(body_html: str) -> list[dict]:
    """Strip HTML and call configured AI provider."""
    if AI_PROVIDER == "none":
        return []
    body_text = _strip_html(body_html)
    body_text = html_lib.unescape(body_text)
    body_text = re.sub(r"\s{3,}", "\n", body_text).strip()
    if AI_PROVIDER == "gemini":
        return _ai_extract_gemini(body_text)
    if AI_PROVIDER == "ollama":
        return _ai_extract_ollama(body_text)
    log.warning(f"Unknown AI_PROVIDER={AI_PROVIDER!r}")
    return []


def _headers_make_sense(col_map: dict) -> bool:
    """
    Return True if at least 3 recognisable candidate fields were mapped.
    Fewer than 3 means the table is probably a layout/nav table, not a
    candidate table — fall through to AI extraction.
    """
    candidate_fields = {
        "name_of_candidate", "contact_number", "email_id",
        "total_experience", "relevant_experience", "current_ctc",
        "expected_ctc", "notice_period", "current_org",
        "current_location", "preferred_location", "qualification",
    }
    mapped = set(col_map.values()) & candidate_fields
    return len(mapped) >= 3


# ── Auth ───────────────────────────────────────────────────────────────────────
def _get_app_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    result = ConfidentialClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
        client_credential=client_secret,
    ).acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(f"App token failed: {result.get('error_description')}")
    return result["access_token"]

def get_mail_token() -> str:
    return _get_app_token(AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET)

def get_onedrive_token() -> str:
    app = PublicClientApplication(
        OD_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{OD_TENANT_ID}",
    )
    result = app.acquire_token_by_refresh_token(
        OD_REFRESH_TOKEN,
        scopes=["https://graph.microsoft.com/.default"],
    )
    if "access_token" not in result:
        raise RuntimeError(f"OneDrive token refresh failed: {result.get('error_description', str(result))}")
    log.info("OneDrive delegated token obtained.")
    return result["access_token"]


# ── Folder resolution ──────────────────────────────────────────────────────────
def resolve_folder_id(token: str, folder_name: str) -> Optional[str]:
    if not folder_name:
        return None
    headers = {"Authorization": f"Bearer {token}"}
    url = (
        f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}/mailFolders"
        f"?$select=id,displayName&$top=50"
    )
    resp = requests.get(url, headers=headers, timeout=15)
    if not resp.ok:
        return None
    for folder in resp.json().get("value", []):
        if folder.get("displayName", "").strip().lower() == folder_name.strip().lower():
            return folder["id"]
        child_url = (
            f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}"
            f"/mailFolders/{folder['id']}/childFolders?$select=id,displayName&$top=50"
        )
        cr = requests.get(child_url, headers=headers, timeout=15)
        if not cr.ok:
            continue
        for child in cr.json().get("value", []):
            if child.get("displayName", "").strip().lower() == folder_name.strip().lower():
                return child["id"]
    log.warning(f"Folder {folder_name!r} not found — using root inbox.")
    return None


# ── Fetch emails ───────────────────────────────────────────────────────────────
def fetch_emails(token: str, folder_id: Optional[str]) -> list[dict]:
    headers = {"Authorization": f"Bearer {token}"}
    since_iso = f"{SINCE_DATE.isoformat()}T00:00:00Z"

    if folder_id:
        base = (
            f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}"
            f"/mailFolders/{folder_id}/messages"
        )
    else:
        base = f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}/messages"

    url = (
        f"{base}"
        f"?$top=50"
        f"&$select=id,subject,from,toRecipients,body,receivedDateTime,"
        f"hasAttachments,conversationId"
        f"&$filter=receivedDateTime ge {since_iso}"
        f"&$orderby=receivedDateTime asc"
    )

    all_emails: list[dict] = []
    page = 0
    while url:
        page += 1
        log.info(f"Fetching email page {page}…")
        try:
            resp = requests.get(url, headers=headers, timeout=90)
            resp.raise_for_status()
        except Exception as exc:
            log.error(f"Error on page {page}: {exc}")
            break
        data = resp.json()
        batch = data.get("value", [])
        all_emails.extend(batch)
        log.info(f"  Page {page}: {len(batch)} email(s) (total: {len(all_emails)})")
        url = data.get("@odata.nextLink")

    log.info(f"Fetched {len(all_emails)} email(s) since {SINCE_DATE}.")
    return all_emails


def fetch_thread_messages(token: str, conversation_id: str) -> list[dict]:
    """
    Fetch all messages in a conversation (thread/reply chain).
    Returns all messages sorted oldest-first.
    """
    headers = {"Authorization": f"Bearer {token}"}
    url = (
        f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}/messages"
        f"?$filter=conversationId eq '{conversation_id}'"
        f"&$select=id,subject,from,toRecipients,body,receivedDateTime,hasAttachments"
        f"&$orderby=receivedDateTime asc"
        f"&$top=20"
    )
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        msgs = resp.json().get("value", [])
        log.info(f"  Thread {conversation_id[:12]}… has {len(msgs)} message(s).")
        return msgs
    except Exception as exc:
        log.warning(f"Failed to fetch thread {conversation_id}: {exc}")
        return []


# ── OneDrive helpers ───────────────────────────────────────────────────────────
def _is_resume(filename: str) -> bool:
    return os.path.splitext(filename)[1].lower() in RESUME_EXTENSIONS

def _fetch_attachment_content(mail_token: str, message_id: str, attachment_id: str) -> bytes:
    url = (
        f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}"
        f"/messages/{message_id}/attachments/{attachment_id}/$value"
    )
    resp = requests.get(url, headers={"Authorization": f"Bearer {mail_token}"}, timeout=120)
    resp.raise_for_status()
    return resp.content

def _ensure_onedrive_folder(od_token: str, folder_name: str) -> None:
    headers = {"Authorization": f"Bearer {od_token}", "Content-Type": "application/json"}
    check = requests.get(
        f"https://graph.microsoft.com/v1.0/me/drive/root:/{folder_name}",
        headers=headers, timeout=30,
    )
    if check.status_code == 200:
        return
    resp = requests.post(
        "https://graph.microsoft.com/v1.0/me/drive/root/children",
        headers=headers,
        json={"name": folder_name, "folder": {}, "@microsoft.graph.conflictBehavior": "rename"},
        timeout=30,
    )
    if resp.status_code in (200, 201):
        log.info(f"Created OneDrive folder '{folder_name}'.")

def _upload_to_onedrive(od_token: str, filename: str, content: bytes) -> Optional[str]:
    remote_path = f"{ONEDRIVE_ARCHIVE_FOLDER}/{filename}" if ONEDRIVE_ARCHIVE_FOLDER else filename
    upload_url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{remote_path}:/content"
    try:
        resp = requests.put(
            upload_url,
            headers={"Authorization": f"Bearer {od_token}", "Content-Type": "application/octet-stream"},
            data=content, timeout=120,
        )
    except requests.exceptions.Timeout:
        log.error(f"Upload timed out for {filename!r}")
        return None
    if resp.status_code not in (200, 201):
        log.error(f"OneDrive upload failed for {filename!r}: {resp.status_code}")
        return None
    item = resp.json()
    item_id = item.get("id")
    try:
        link_resp = requests.post(
            f"https://graph.microsoft.com/v1.0/me/drive/items/{item_id}/createLink",
            headers={"Authorization": f"Bearer {od_token}", "Content-Type": "application/json"},
            json={"type": "view", "scope": "organization"}, timeout=30,
        )
        if link_resp.status_code in (200, 201):
            web_url = link_resp.json().get("link", {}).get("webUrl", "")
            if web_url:
                log.info(f"  ↑ Uploaded {filename!r} → {web_url}")
                return web_url
    except Exception as exc:
        log.warning(f"createLink failed: {exc}")
    return item.get("webUrl", "")

def upload_attachments(od_token: str, mail_token: str, msg: dict) -> str:
    if SKIP_UPLOADS or not msg.get("hasAttachments"):
        return ""
    url = (
        f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}"
        f"/messages/{msg['id']}/attachments?$select=id,name,contentType"
    )
    try:
        resp = requests.get(url, headers={"Authorization": f"Bearer {mail_token}"}, timeout=60)
        resp.raise_for_status()
        attachments = resp.json().get("value", [])
    except Exception as exc:
        log.warning(f"Could not fetch attachments: {exc}")
        return ""
    links: list[str] = []
    for att in attachments:
        filename = att.get("name", "")
        if not filename or not _is_resume(filename):
            continue
        att_id = att.get("id", "")
        try:
            content = _fetch_attachment_content(mail_token, msg["id"], att_id)
            link = _upload_to_onedrive(od_token, filename, content)
            links.append(link if link else filename)
        except Exception as exc:
            log.error(f"Failed to upload {filename!r}: {exc}")
            links.append(filename)
    return ", ".join(links)


# ── Subject filter ─────────────────────────────────────────────────────────────
SKIP_SUBJECT_RE = re.compile(
    r"\b(undelivered|undeliverable|delivery\s+failed|delivery\s+status"
    r"|delivery\s+notification|mail\s+delivery|returned\s+mail"
    r"|non[- ]?deliverable|bounced?|failed\s+delivery|delivered)\b",
    re.IGNORECASE,
)

SUBJECT_RE = re.compile(
    r"(?<!\S)(?:\[External\]\s*:\s*)?(?P<code>[A-Za-z]{2,4})\s*:",
    re.IGNORECASE,
)

_JR_PATTERNS = [
    re.compile(r"[-|]\s*(?:jr\s*)?(?P<jr>\d{4,})\s*$",   re.IGNORECASE),
    re.compile(r"^\s*(?:jr\s*)?(?P<jr>\d{4,})\s*[-|]",   re.IGNORECASE),
    re.compile(r"\(\s*jr\s*(?P<jr>\d{4,})\s*\)",          re.IGNORECASE),
    re.compile(r"\(\s*(?P<jr>\d{4,})\s*\)",               re.IGNORECASE),
    re.compile(r"\bjr\s*(?P<jr>\d{4,})\b",                re.IGNORECASE),
    re.compile(r"\b(?P<jr>\d{4,})\b",                     re.IGNORECASE),
]

def _extract_jr_and_skill(rest: str) -> tuple[Optional[str], str]:
    rest = re.sub(r"^[-|%\s]+|[-|%\s]+$", "", rest.strip()).strip()
    jr_no = None
    for pat in _JR_PATTERNS:
        m = pat.search(rest)
        if m:
            jr_no = m.group("jr")
            s, e = m.span()
            rest = re.sub(r"^[-|%\s]+|[-|%\s]+$", "", (rest[:s] + " " + rest[e:]).strip()).strip()
            break
    return jr_no, re.sub(r"\s{2,}", " ", rest).strip()

def parse_subject(subject: str) -> Optional[tuple[str, str, str, Optional[str]]]:
    subject = subject.strip()
    if SKIP_SUBJECT_RE.search(subject):
        return None
    m = SUBJECT_RE.search(subject)
    if not m:
        return None
    code = m.group("code").upper()
    rest = subject[m.end():].strip()
    company_name = COMPANY_CODES.get(code, code)
    jr_no, skill = _extract_jr_and_skill(rest)
    return code, company_name, skill, jr_no


# ── HTML table parser ──────────────────────────────────────────────────────────
def _strip_html(text: str) -> str:
    return re.sub(r"<[^>]+>", "", text)

def _clean_cell(text: str) -> str:
    text = html_lib.unescape(text)
    text = _strip_html(text)
    text = text.replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()

# ── Sanitisers ─────────────────────────────────────────────────────────────────
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]{2,}$")
_PHONE_RE = re.compile(r"^[\d\s\-\+\(\)]{7,15}$")

def _looks_like_email(val: str) -> bool:
    return bool(val and _EMAIL_RE.match(val.strip()))

def _looks_like_phone(val: str) -> bool:
    return bool(val and _PHONE_RE.match(val.strip()) and "@" not in val)

def _sanitize_contact(val: Optional[str]) -> Optional[str]:
    if not val:
        return None
    v = val.strip()
    if "@" in v:
        return None
    digits = re.sub(r"[^\d+]", "", v)
    digits = re.sub(r"(?<!^)\+", "", digits)
    return digits if len(digits) >= 7 else None

def _sanitize_email(val: Optional[str]) -> Optional[str]:
    if not val:
        return None
    v = html_lib.unescape(val).strip()
    v = re.sub(r"^[^\w]+", "", v).strip().lower()
    return v if _EMAIL_RE.match(v) else None

def _fix_swapped_contact_email(record: dict) -> dict:
    contact = (record.get("contact_number") or "").strip()
    email   = (record.get("email_id") or "").strip()
    if _looks_like_email(contact) and _looks_like_phone(email):
        record["contact_number"], record["email_id"] = email, contact
    elif _looks_like_email(contact) and not email:
        record["email_id"] = contact
        record["contact_number"] = None
    elif _looks_like_phone(email) and not contact:
        record["contact_number"] = email
        record["email_id"] = None
    return record


# ── HTML table parser with header-sense check ──────────────────────────────────
def parse_html_table(html: str) -> tuple[list[dict], bool]:
    """
    Returns (rows, headers_made_sense).
    headers_made_sense=False means AI fallback should be triggered.
    """
    rows: list[dict] = []
    any_table_found = False
    any_sensible = False

    for tbl in re.findall(r"<table[^>]*>(.*?)</table>", html, re.DOTALL | re.IGNORECASE):
        all_rows = re.findall(r"<tr[^>]*>(.*?)</tr>", tbl, re.DOTALL | re.IGNORECASE)
        if not all_rows:
            continue
        any_table_found = True
        raw_headers = re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", all_rows[0], re.DOTALL | re.IGNORECASE)
        headers = [_clean_cell(h) for h in raw_headers]
        col_map = {i: _resolve_header(h) for i, h in enumerate(headers) if _resolve_header(h)}

        if not _headers_make_sense(col_map):
            log.info(f"  Table found but headers not sensible ({headers[:5]}) — will try AI.")
            continue

        any_sensible = True
        for row_html in all_rows[1:]:
            cells = re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", row_html, re.DOTALL | re.IGNORECASE)
            if not cells:
                continue
            cleaned = [_clean_cell(c) for c in cells]
            if all(v == "" for v in cleaned):
                continue
            record = {col_map[i]: v for i, v in enumerate(cleaned) if i in col_map}
            if record:
                rows.append(record)

    # Need AI if: no table at all, OR tables exist but none had sensible headers
    need_ai = (not rows) or (any_table_found and not any_sensible)
    return rows, not need_ai


# ── Misc helpers ───────────────────────────────────────────────────────────────
def _extract_address(addr_obj: dict) -> str:
    try:
        return addr_obj["emailAddress"]["address"].strip().lower()
    except (KeyError, TypeError):
        return ""

def _normalize_name_from_email(email: str) -> str:
    if not email or "@" not in email:
        return ""
    local = re.sub(r"\d+", "", email.split("@", 1)[0])
    local = re.sub(r"[._\-]+", " ", local)
    return re.sub(r"\s+", " ", local).strip().title()

_DATE_FORMATS = ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y",
                 "%d-%b-%y", "%d %b %Y", "%d-%b-%Y")

def _parse_date(raw: str) -> Optional[date]:
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw.strip(), fmt).date()
        except ValueError:
            continue
    return None

_REQUIRED_FIELDS = (
    "name_of_candidate", "contact_number", "email_id",
    "general_skill", "company_name", "recruiter", "email_from", "email_to",
)

def _record_status(data: dict) -> str:
    missing = [f for f in _REQUIRED_FIELDS if not data.get(f)]
    return "Fail" if missing else "Pass"

def _t(val) -> Optional[str]:
    if val is None:
        return None
    v = str(val).strip()
    return v or None


# ── DB helpers ─────────────────────────────────────────────────────────────────
def _make_insert_sql() -> pgsql.Composed:
    return pgsql.SQL(
        "INSERT INTO {schema}.{table} ({fields}) VALUES ({placeholders})"
    ).format(
        schema=pgsql.Identifier(DB_SCHEMA),
        table=pgsql.Identifier(DB_TABLE),
        fields=pgsql.SQL(", ").join(map(pgsql.Identifier, DB_COLUMNS)),
        placeholders=pgsql.SQL(", ").join(pgsql.Placeholder() * len(DB_COLUMNS)),
    )

def _verify_table_columns(conn) -> None:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_schema = %s AND table_name = %s",
            (DB_SCHEMA, DB_TABLE),
        )
        existing = {row[0] for row in cur.fetchall()}
    if not existing:
        log.error(f"Table {DB_SCHEMA}.{DB_TABLE} not found.")
        raise SystemExit(1)
    missing = [c for c in DB_COLUMNS if c not in existing]
    if missing:
        log.error(f"Columns missing: {missing}")
        raise SystemExit(1)
    log.info(f"Column pre-flight passed for {DB_SCHEMA}.{DB_TABLE}.")

def insert_records_incremental(conn, records: list[dict]) -> int:
    insert_sql = _make_insert_sql()
    inserted = 0
    for rec in records:
        row = tuple(rec.get(col) for col in DB_COLUMNS)
        try:
            with conn.cursor() as cur:
                cur.execute(insert_sql, row)
            conn.commit()
            inserted += 1
        except Exception as exc:
            log.error(f"Row insert failed for {rec.get('name_of_candidate', '?')!r}: {exc}")
            conn.rollback()
    return inserted


# ── Excel writer ───────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
AI_FILL     = PatternFill("solid", start_color="FFF2CC")   # yellow tint for AI rows

COL_WIDTHS = {
    "email_subject": 40, "recruiter": 18, "client_recruiter": 18,
    "email_from": 28, "email_to": 28, "delivery_type": 14,
    "company_name": 16, "general_skill": 22, "date": 12, "jr_no": 10,
    "name_of_candidate": 22, "contact_number": 16, "email_id": 28,
    "total_experience": 14, "relevant_experience": 16,
    "current_ctc": 14, "expected_ctc": 14, "notice_period": 14,
    "current_org": 22, "current_location": 18, "preferred_location": 18,
    "qualification": 18, "attachment": 35, "remarks": 30,
    "record_status": 14, "ai_extracted": 12,
}

PRETTY_HEADERS = {
    "email_subject": "Email Subject", "recruiter": "Recruiter",
    "client_recruiter": "Client Recruiter", "email_from": "Email From",
    "email_to": "Email To", "delivery_type": "Delivery Type",
    "company_name": "Company", "general_skill": "General Skill",
    "date": "Date", "jr_no": "JR No",
    "name_of_candidate": "Candidate Name", "contact_number": "Contact Number",
    "email_id": "Email ID", "total_experience": "Total Exp",
    "relevant_experience": "Relevant Exp", "current_ctc": "Current CTC",
    "expected_ctc": "Expected CTC", "notice_period": "Notice Period",
    "current_org": "Current Org", "current_location": "Current Location",
    "preferred_location": "Preferred Location", "qualification": "Qualification",
    "attachment": "Attachment", "remarks": "Remarks",
    "record_status": "Record Status", "ai_extracted": "AI Extracted",
}

STATUS_COLORS = {"Pass": "C6EFCE", "Fail": "FFC7CE"}

def write_excel(records: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"
    ws.freeze_panes = "A2"

    for col_idx, col in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=PRETTY_HEADERS.get(col, col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = COL_WIDTHS.get(col, 16)
    ws.row_dimensions[1].height = 22

    for row_idx, rec in enumerate(records, 2):
        is_ai = rec.get("ai_extracted") == "Yes"
        for col_idx, col in enumerate(EXCEL_COLUMNS, 1):
            val  = rec.get(col)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = LEFT
            if col == "record_status" and val in STATUS_COLORS:
                cell.fill = PatternFill("solid", start_color=STATUS_COLORS[val])
            elif is_ai and col not in ("record_status",):
                cell.fill = AI_FILL
        ws.row_dimensions[row_idx].height = 16

    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    log.info(f"Saved {len(records)} row(s) → {output_path}")


# ── Build candidate record ─────────────────────────────────────────────────────
def _build_record(
    row: dict,
    subject: str,
    recruiter: str,
    client_recruiter: str,
    from_addr: str,
    to_addr: str,
    delivery_type: str,
    company_name: str,
    general_skill: str,
    subject_jr_no: Optional[str],
    email_date: Optional[date],
    attachment_str: str,
    ai_extracted: bool = False,
) -> dict:
    row_date = _parse_date(row["date"]) if row.get("date") else None
    if row_date is None:
        row_date = email_date or date.today()

    effective_jr_no = _t(row.get("jr_no")) or subject_jr_no
    contact_number  = _sanitize_contact(_t(row.get("contact_number")))
    email_id_val    = _sanitize_email(_t(row.get("email_id")))

    record = {
        "email_subject":       subject,
        "recruiter":           recruiter,
        "client_recruiter":    client_recruiter,
        "email_from":          from_addr,
        "email_to":            to_addr,
        "delivery_type":       delivery_type,
        "company_name":        company_name,
        "general_skill":       _t(general_skill),
        "date":                row_date,
        "jr_no":               effective_jr_no,
        "name_of_candidate":   _t(row.get("name_of_candidate")),
        "contact_number":      contact_number,
        "email_id":            email_id_val,
        "total_experience":    _t(row.get("total_experience")),
        "relevant_experience": _t(row.get("relevant_experience")),
        "current_ctc":         _t(row.get("current_ctc")),
        "expected_ctc":        _t(row.get("expected_ctc")),
        "notice_period":       _t(row.get("notice_period")),
        "current_org":         _t(row.get("current_org")),
        "current_location":    _t(row.get("current_location")),
        "preferred_location":  _t(row.get("preferred_location")),
        "qualification":       _t(row.get("qualification")),
        "attachment":          _t(attachment_str),
        "remarks":             _t(row.get("remarks")),
        "ai_extracted":        "Yes" if ai_extracted else "No",
    }

    record = _fix_swapped_contact_email(record)

    record["record_status"] = _record_status({
        "name_of_candidate": record.get("name_of_candidate"),
        "contact_number":    record.get("contact_number"),
        "email_id":          record.get("email_id"),
        "general_skill":     _t(general_skill),
        "company_name":      company_name,
        "recruiter":         recruiter,
        "email_from":        from_addr,
        "email_to":          to_addr,
    })

    return record


# ── Process a single message ───────────────────────────────────────────────────
def process_message(
    msg: dict,
    subject: str,
    company_name: str,
    general_skill: str,
    subject_jr_no: Optional[str],
    recruiter: str,
    client_recruiter: str,
    from_addr: str,
    to_addr: str,
    delivery_type: str,
    od_token: str,
    mail_token: str,
) -> list[dict]:
    """
    Parse one message: try HTML table first, fall back to AI if needed.
    Returns list of candidate records extracted from this message.
    """
    body_html = (msg.get("body") or {}).get("content", "")
    attachment_str = upload_attachments(od_token, mail_token, msg)

    email_date: Optional[date] = None
    raw_dt = msg.get("receivedDateTime", "")
    if raw_dt:
        try:
            email_date = datetime.fromisoformat(raw_dt.replace("Z", "+00:00")).date()
        except ValueError:
            pass

    rows, headers_ok = parse_html_table(body_html)

    use_ai = not headers_ok or not rows

    records: list[dict] = []

    if rows and headers_ok:
        for row in rows:
            rec = _build_record(
                row, subject, recruiter, client_recruiter, from_addr, to_addr,
                delivery_type, company_name, general_skill, subject_jr_no,
                email_date, attachment_str, ai_extracted=False,
            )
            records.append(rec)
        log.info(f"  Table parsed: {len(rows)} candidate(s) from message.")

    if use_ai and AI_PROVIDER != "none":
        log.info(f"  Triggering AI extraction for message (headers_ok={headers_ok}, rows={len(rows)}).")
        ai_rows = ai_extract(body_html)
        for row in ai_rows:
            rec = _build_record(
                row, subject, recruiter, client_recruiter, from_addr, to_addr,
                delivery_type, company_name, general_skill, subject_jr_no,
                email_date, attachment_str, ai_extracted=True,
            )
            records.append(rec)
        if not ai_rows:
            log.warning(f"  AI returned no candidates for: {subject!r}")
            if not records:
                # Add skeleton so the email is at least visible in the report
                rec = _build_record(
                    {}, subject, recruiter, client_recruiter, from_addr, to_addr,
                    delivery_type, company_name, general_skill, subject_jr_no,
                    email_date, attachment_str, ai_extracted=True,
                )
                records.append(rec)
    elif use_ai and AI_PROVIDER == "none" and not records:
        log.warning(f"  No table & AI disabled — skeleton row for: {subject!r}")
        rec = _build_record(
            {}, subject, recruiter, client_recruiter, from_addr, to_addr,
            delivery_type, company_name, general_skill, subject_jr_no,
            email_date, attachment_str, ai_extracted=False,
        )
        records.append(rec)

    return records


# ── Main ───────────────────────────────────────────────────────────────────────
def run() -> None:
    log.info("=== Ad-hoc extractor v2 starting ===")
    log.info(f"Target mailbox : {TARGET_MAILBOX}")
    log.info(f"OneDrive user  : {ONEDRIVE_USER}")
    log.info(f"Pulling from   : {SINCE_DATE} onwards")
    log.info(f"Folder         : {INBOX_SUBFOLDER or '(root inbox)'}")
    log.info(f"AI provider    : {AI_PROVIDER}")
    log.info(f"Only complete  : {ONLY_COMPLETE}")
    log.info(f"Skip uploads   : {SKIP_UPLOADS}")
    log.info(f"DB target      : {DB_SCHEMA}.{DB_TABLE}")

    mail_token = get_mail_token()
    od_token   = get_onedrive_token()
    log.info("Tokens obtained.")

    if not SKIP_UPLOADS:
        check = requests.get(
            "https://graph.microsoft.com/v1.0/me/drive",
            headers={"Authorization": f"Bearer {od_token}"},
            timeout=30,
        )
        if check.status_code == 200:
            log.info(f"OneDrive access confirmed.")
            if ONEDRIVE_ARCHIVE_FOLDER:
                _ensure_onedrive_folder(od_token, ONEDRIVE_ARCHIVE_FOLDER)
        else:
            log.warning(f"OneDrive check returned {check.status_code} — uploads may fail.")

    try:
        db_conn = psycopg2.connect(DB_DSN)
        db_conn.autocommit = False
        log.info("DB connection opened.")
        _verify_table_columns(db_conn)
    except SystemExit:
        raise
    except Exception as exc:
        log.error(f"Cannot connect to DB: {exc}")
        db_conn = None

    folder_id = resolve_folder_id(mail_token, INBOX_SUBFOLDER) if INBOX_SUBFOLDER else None
    emails    = fetch_emails(mail_token, folder_id)

    # Deduplicate: track processed conversation IDs so we don't process
    # both the original email AND its replies as separate top-level emails.
    processed_conversations: set[str] = set()

    all_records: list[dict] = []
    skipped  = 0
    db_count = 0

    for msg in emails:
        subject = msg.get("subject", "").strip()
        parsed  = parse_subject(subject)
        if parsed is None:
            skipped += 1
            continue

        company_code, company_name, general_skill, subject_jr_no = parsed
        conversation_id = msg.get("conversationId", "")

        from_addr = _extract_address(msg.get("from", {}))
        to_list   = msg.get("toRecipients", [])
        to_addr   = _extract_address(to_list[0]) if to_list else ""

        recruiter        = _normalize_name_from_email(from_addr)
        client_recruiter = _normalize_name_from_email(to_addr)
        delivery_type    = (
            "Internal"
            if VOLIBITS_DOMAIN in from_addr and VOLIBITS_DOMAIN in to_addr
            else "External"
        )

        # ── Collect all messages in the thread ────────────────────────────
        if conversation_id and conversation_id not in processed_conversations:
            thread_msgs = fetch_thread_messages(mail_token, conversation_id)
            processed_conversations.add(conversation_id)
        else:
            # Already processed this conversation via an earlier email
            if conversation_id in processed_conversations:
                log.info(f"Thread already processed, skipping: {subject!r}")
                skipped += 1
                continue
            thread_msgs = [msg]

        if not thread_msgs:
            thread_msgs = [msg]

        email_records: list[dict] = []

        for thread_msg in thread_msgs:
            recs = process_message(
                thread_msg, subject, company_name, general_skill, subject_jr_no,
                recruiter, client_recruiter, from_addr, to_addr, delivery_type,
                od_token, mail_token,
            )
            for rec in recs:
                cand = rec.get("name_of_candidate") or "(unknown)"
                log.info(f"  + {cand} | {subject!r} | {rec.get('date')} | AI={rec.get('ai_extracted')}")
            email_records.extend(recs)

        # Deduplicate candidates within the thread by (name, contact)
        seen_candidates: set[tuple] = set()
        deduped: list[dict] = []
        for rec in email_records:
            key = (
                (rec.get("name_of_candidate") or "").lower().strip(),
                (rec.get("contact_number") or "").strip(),
            )
            if key == ("", ""):
                deduped.append(rec)  # skeleton rows always included
            elif key not in seen_candidates:
                seen_candidates.add(key)
                deduped.append(rec)
            else:
                log.info(f"  Duplicate candidate skipped: {key}")

        # Apply ONLY_COMPLETE filter if requested
        if ONLY_COMPLETE:
            before = len(deduped)
            deduped = [r for r in deduped if r.get("record_status") == "Pass"]
            if len(deduped) < before:
                log.info(f"  ONLY_COMPLETE: dropped {before - len(deduped)} incomplete row(s).")

        all_records.extend(deduped)

        if deduped and db_conn is not None:
            n = insert_records_incremental(db_conn, deduped)
            db_count += n
            log.info(f"  DB committed {n}/{len(deduped)} row(s) (running total: {db_count})")

    if db_conn is not None:
        try:
            db_conn.close()
            log.info("DB connection closed.")
        except Exception:
            pass

    log.info(f"Skipped  : {skipped} email(s)")
    log.info(f"Total rows: {len(all_records)}")
    log.info(f"DB rows  : {db_count}")

    if not all_records:
        log.warning("No matching records found.")
        return

    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"adhoc_candidates_{ts}.xlsx"
    write_excel(all_records, output_path)

    pass_count = sum(1 for r in all_records if r.get("record_status") == "Pass")
    ai_count   = sum(1 for r in all_records if r.get("ai_extracted") == "Yes")

    print(
        f"\n✅  Done — {len(all_records)} row(s) processed.\n"
        f"    Pass / Fail      : {pass_count} / {len(all_records) - pass_count}\n"
        f"    AI-extracted rows: {ai_count}\n"
        f"    DB rows inserted : {db_count}\n"
        f"    Excel output     : {output_path}\n"
    )


if __name__ == "__main__":
    run()