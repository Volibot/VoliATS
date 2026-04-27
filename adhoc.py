"""
Ad-hoc extractor: pulls candidate emails from "Company Profiles" subfolder
received on or after 2026-01-01, writes parsed fields to the
`public.temp_hrvolibit_archive` PostgreSQL table, and saves an Excel report.

Subject filters accepted:
  - CODE: ...
  - [External]: CODE: ...

Attachments are uploaded to an Archive folder in OneDrive.

Output: adhoc_candidates_<timestamp>.xlsx in the current working directory.

Required env vars:
  AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET
  TARGET_MAILBOX
  OD_TENANT_ID, OD_CLIENT_ID, OD_REFRESH_TOKEN, ONEDRIVE_USER
  DB_DSN

Optional:
  OD_CLIENT_SECRET        (OneDrive secret — can be empty for public apps)
  INBOX_SUBFOLDER         (default: "Company Profiles")
  DB_SCHEMA               (default: "public")
  DB_TABLE_NAME           (default: "temp_hrvolibit_archive")
  ONEDRIVE_ARCHIVE_FOLDER (default: "archive")
  LIMIT                   (default: 100)
  SKIP_UPLOADS            (default: false — set "true" to skip OneDrive uploads)

── Local run ────────────────────────────────────────────────────────────────────
1. Install dependencies:
       pip install msal openpyxl requests python-dotenv psycopg2-binary

2. Create a .env file next to this script (⚠️ never commit it to git):

       AZURE_TENANT_ID=xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx
       AZURE_CLIENT_ID=xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx
       AZURE_CLIENT_SECRET=your-secret-here
       TARGET_MAILBOX=mailbox@yourdomain.com
       OD_TENANT_ID=xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx
       OD_CLIENT_ID=xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx
       OD_REFRESH_TOKEN=your-refresh-token
       ONEDRIVE_USER=user@yourdomain.com
       DB_DSN=postgresql://user:password@host:5432/dbname

3. Run:
       python bs_extractor.py
────────────────────────────────────────────────────────────────────────────────
"""

import os
import re
import html as html_lib
import logging
import requests
import psycopg2
from psycopg2 import sql as pgsql
from datetime import datetime, date
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from msal import ConfidentialClientApplication, PublicClientApplication
from dotenv import load_dotenv

load_dotenv()

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler("adhoc_extract.log")],
)
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
AZURE_TENANT_ID     = os.environ["AZURE_TENANT_ID"]
AZURE_CLIENT_ID     = os.environ["AZURE_CLIENT_ID"]
AZURE_CLIENT_SECRET = os.environ["AZURE_CLIENT_SECRET"]

TARGET_MAILBOX  = os.environ["TARGET_MAILBOX"]
INBOX_SUBFOLDER = os.environ.get("INBOX_SUBFOLDER", "Company Profiles")
VOLIBITS_DOMAIN = "volibits.com"
RESUME_EXTENSIONS = {".pdf", ".doc", ".docx"}

OD_TENANT_ID            = os.environ["OD_TENANT_ID"]
OD_CLIENT_ID            = os.environ["OD_CLIENT_ID"]
OD_CLIENT_SECRET        = os.environ.get("OD_CLIENT_SECRET", "")
OD_REFRESH_TOKEN        = os.environ["OD_REFRESH_TOKEN"]
ONEDRIVE_USER           = os.environ["ONEDRIVE_USER"]
DB_DSN                  = os.environ["DB_DSN"]

DB_SCHEMA = os.environ.get("DB_SCHEMA", "public")
DB_TABLE  = "temp_hrvolibit_archive"

ONEDRIVE_ARCHIVE_FOLDER = os.environ.get("ONEDRIVE_ARCHIVE_FOLDER", "archive")
LIMIT                   = int(os.environ.get("LIMIT", "100"))

SKIP_UPLOADS = os.environ.get("SKIP_UPLOADS", "false")

COMPANY_CODES: dict[str, str] = {
    "BS":  "Birlasoft",
    "BW":  "BeWealthy",
    "TCS": "TCS",
    "INF": "Infosys",
    "WIP": "Wipro",
    "HCL": "HCL Technologies",
    "ACC": "Accenture",
    "CAP": "Capgemini",
    "COG": "Cognizant",
    "IBM": "IBM",
    "SAP": "SAP",
    "ORC": "Oracle",
    "MS":  "Microsoft",
    "RS":  "RS Software",
}

SINCE_DATE = date(2026, 1, 1)

EXCEL_COLUMNS = [
    "email_subject",
    "recruiter", "client_recruiter", "email_from", "email_to",
    "delivery_type", "company_name", "general_skill",
    "date", "jr_no",
    "name_of_candidate", "contact_number", "email_id",
    "total_experience", "relevant_experience",
    "current_ctc", "expected_ctc", "notice_period",
    "current_org", "current_location", "preferred_location",
    "attachment", "remarks", "record_status",
]

DB_COLUMNS = [
    "email_subject",
    "recruiter",
    "client_recruiter",
    "email_from",
    "email_to",
    "delivery_type",
    "company_name",
    "general_skill",
    "date",
    "jr_no",
    "name_of_candidate",
    "contact_number",
    "email_id",
    "total_experience",
    "relevant_experience",
    "current_ctc",
    "expected_ctc",
    "notice_period",
    "current_org",
    "current_location",
    "preferred_location",
    "attachment",
    "remarks",
    "record_status",
]

# ── Column aliases ─────────────────────────────────────────────────────────────
COLUMN_ALIASES: dict[str, list[str]] = {
    "general_skill": [
        "general_skill", "generalskill", "gen_skill", "genskill",
        "skill", "requirement",
    ],
    "jr_no": [
        "jr no", "jr_no", "jr no.", "jr", "jrno", "req_id", "req id",
        "requisition id", "requisition_id", "jr_number", "jr number",
        "job req", "job requisition", "job_req", "reqid",
        "requirement id", "requirement_id",
        "jr no(mention the jr number where profiles are uploaded on sf).",
        "jr_no.", "(mention the jr number where profiles are uploaded on sf).",
    ],
    "name_of_candidate": [
        "candidate name", "candidate_name", "name", "applicant name",
        "applicant_name", "full name", "full_name", "name of candidate",
        "candidate", "person name",
        "n_of_candidate", "noc", "name of cand", "name_of_cand",
        "name_of_candidate",
    ],
    "contact_number": [
        "contact number", "contact_number", "phone", "mobile", "cell",
        "phone number", "phone_number", "mobile number", "mobile_number",
        "contact no", "contact_no", "ph no", "ph_no", "contact",
        "c_number", "con_number", "con_no.", "con_num", "con no.", "cno.",
        "mobile no.", "mobile no", "contact no", "contact number",
    ],
    "email_id": [
        "email id", "email_id", "email", "e-mail", "mail id", "mail_id",
        "email address", "email_address", "e mail",
        "e_id", "e id", "emailid", "e-mail id",
    ],
    "current_org": [
        "current company", "current_company", "company", "employer",
        "current employer", "current_employer", "current org", "current_org",
        "organisation", "organization", "curr company", "present company",
        "curr_org", "current_organization", "current comp", "curr comp",
        "current / last company",
    ],
    "total_experience": [
        "total experience", "total_experience", "total exp", "total_exp",
        "experience", "exp", "total yrs", "total years",
        "tot exp", "tot_exp", "total exp (yrs)",
    ],
    "relevant_experience": [
        "relevant experience", "relevant_experience", "relevant exp",
        "relevant_exp", "rel exp", "rel_exp", "relevant yrs",
        "rel exp.", "rel experience", "rel. exp", "rel.exp",
        "rel_experience", "rel.experience", "relevant exp (yrs)",
    ],
    "current_ctc": [
        "current ctc", "current_ctc", "ctc", "current salary",
        "current_salary", "present ctc", "present_ctc", "curr ctc",
        "cctc", "currctc", "curr_ctc", "c_ctc",
        "current ctc (lakhs)", "ctcc",
    ],
    "expected_ctc": [
        "expected ctc", "expected_ctc", "expected salary", "expected_salary",
        "exp ctc", "exp_ctc", "desired ctc", "desired_ctc",
        "ectc", "e_ctc", "epected ctc", "expctc", "ect",
        "exp ctc (lakhs)", "exp ctc",
    ],
    "notice_period": [
        "notice period", "notice_period", "notice", "np", "joining time",
        "notice days", "notice_days",
        "notice period(days)",
    ],
    "current_location": [
        "current location", "current_location", "location", "city",
        "current city", "current_city", "present location",
        "current_loc", "curr location", "c loc", "c_location", "curr_location",
        "curr_loc", "current/preferred location",
    ],
    "preferred_location": [
        "preferred location", "preferred_location", "preferred city",
        "preferred_city", "pref location", "pref_location",
        "willing to relocate", "target location",
        "pref_loc", "pre location", "pref_location",
        "preferred location", "current/preferred location",
    ],
    "date": [
        "date", "submission date", "submission_date", "applied date",
        "applied_date", "profile date",
        "date of submission",
    ],
    "remarks": [
        "remarks", "comments", "comment", "note", "notes",
        "remarks/availability for interview",
    ],
}

def _normalize(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())

ALIAS_MAP: dict[str, str] = {
    _normalize(alias): canonical
    for canonical, aliases in COLUMN_ALIASES.items()
    for alias in aliases
}

def _resolve_header(raw: str) -> Optional[str]:
    return ALIAS_MAP.get(_normalize(raw))


# ── Auth ───────────────────────────────────────────────────────────────────────
def _get_app_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    result = ConfidentialClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
        client_credential=client_secret,
    ).acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(
            f"App token acquisition failed: {result.get('error_description')}"
        )
    return result["access_token"]


def get_mail_token() -> str:
    return _get_app_token(AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET)


def get_onedrive_token() -> str:
    app = PublicClientApplication(
        OD_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{OD_TENANT_ID}",
    )
    result = app.acquire_token_by_refresh_token(
        OD_REFRESH_TOKEN,
        scopes=["https://graph.microsoft.com/.default"],
    )
    if "access_token" not in result:
        raise RuntimeError(
            f"OneDrive token refresh failed: {result.get('error_description', str(result))}\n"
            "Re-run generate_refresh_token.py and update OD_REFRESH_TOKEN."
        )
    log.info("OneDrive delegated token obtained successfully.")
    return result["access_token"]


# ── Folder resolution ──────────────────────────────────────────────────────────
def resolve_folder_id(token: str, folder_name: str) -> Optional[str]:
    if not folder_name:
        return None
    headers = {"Authorization": f"Bearer {token}"}
    url = (
        f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}/mailFolders"
        f"?$select=id,displayName&$top=50"
    )
    resp = requests.get(url, headers=headers, timeout=15)
    if not resp.ok:
        return None
    for folder in resp.json().get("value", []):
        if folder.get("displayName", "").strip().lower() == folder_name.strip().lower():
            return folder["id"]
        child_url = (
            f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}"
            f"/mailFolders/{folder['id']}/childFolders?$select=id,displayName&$top=50"
        )
        cr = requests.get(child_url, headers=headers, timeout=15)
        if not cr.ok:
            continue
        for child in cr.json().get("value", []):
            if child.get("displayName", "").strip().lower() == folder_name.strip().lower():
                return child["id"]
    log.warning(f"Folder {folder_name!r} not found — will use root inbox")
    return None


# ── Fetch emails ───────────────────────────────────────────────────────────────
def fetch_bs_emails(token: str, folder_id: Optional[str]) -> list[dict]:
    headers = {"Authorization": f"Bearer {token}"}
    since_iso = f"{SINCE_DATE.isoformat()}T00:00:00Z"

    if folder_id:
        base = (
            f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}"
            f"/mailFolders/{folder_id}/messages"
        )
    else:
        base = f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}/messages"

    url = (
        f"{base}"
        f"?$top=50"
        f"&$select=id,subject,from,toRecipients,body,receivedDateTime,hasAttachments"
        f"&$filter=receivedDateTime ge {since_iso}"
        f"&$orderby=receivedDateTime asc"
    )

    all_emails: list[dict] = []
    page = 0
    while url:
        page += 1
        log.info(f"Fetching email page {page}…")
        try:
            resp = requests.get(url, headers=headers, timeout=90)
            resp.raise_for_status()
        except requests.exceptions.Timeout:
            log.error(f"Timeout on page {page} — stopping. Got {len(all_emails)} emails so far.")
            break
        except requests.exceptions.RequestException as exc:
            log.error(f"Request error on page {page}: {exc} — stopping.")
            break

        data = resp.json()
        batch = data.get("value", [])
        all_emails.extend(batch)
        log.info(f"  Page {page}: {len(batch)} email(s) (total so far: {len(all_emails)})")
        url = data.get("@odata.nextLink")

    log.info(f"Fetched {len(all_emails)} email(s) since {SINCE_DATE}.")
    return all_emails


# ── OneDrive attachment helpers ────────────────────────────────────────────────
def _is_resume(filename: str) -> bool:
    return os.path.splitext(filename)[1].lower() in RESUME_EXTENSIONS


def _fetch_attachment_content(mail_token: str, message_id: str, attachment_id: str) -> bytes:
    url = (
        f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}"
        f"/messages/{message_id}/attachments/{attachment_id}/$value"
    )
    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {mail_token}"},
        timeout=120,
    )
    resp.raise_for_status()
    return resp.content


def _ensure_onedrive_folder(od_token: str, folder_name: str) -> None:
    headers = {
        "Authorization": f"Bearer {od_token}",
        "Content-Type": "application/json",
    }
    check = requests.get(
        f"https://graph.microsoft.com/v1.0/me/drive/root:/{folder_name}",
        headers=headers,
        timeout=30,
    )
    if check.status_code == 200:
        log.info(f"OneDrive folder '{folder_name}' already exists.")
        return
    resp = requests.post(
        "https://graph.microsoft.com/v1.0/me/drive/root/children",
        headers=headers,
        json={
            "name": folder_name,
            "folder": {},
            "@microsoft.graph.conflictBehavior": "rename",
        },
        timeout=30,
    )
    if resp.status_code in (200, 201):
        log.info(f"Created OneDrive folder '{folder_name}'.")
    else:
        log.warning(
            f"Could not create OneDrive folder '{folder_name}': "
            f"{resp.status_code} — {resp.text[:200]}"
        )


def _upload_to_onedrive(od_token: str, filename: str, content: bytes) -> Optional[str]:
    remote_path = f"{ONEDRIVE_ARCHIVE_FOLDER}/{filename}" if ONEDRIVE_ARCHIVE_FOLDER else filename
    upload_url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{remote_path}:/content"
    try:
        resp = requests.put(
            upload_url,
            headers={
                "Authorization": f"Bearer {od_token}",
                "Content-Type": "application/octet-stream",
            },
            data=content,
            timeout=120,
        )
    except requests.exceptions.Timeout:
        log.error(f"Upload timed out for {filename!r}")
        return None

    if resp.status_code not in (200, 201):
        log.error(
            f"OneDrive upload failed for {filename!r}: "
            f"{resp.status_code} — {resp.text[:300]}"
        )
        return None

    item = resp.json()
    item_id = item.get("id")
    try:
        link_resp = requests.post(
            f"https://graph.microsoft.com/v1.0/me/drive/items/{item_id}/createLink",
            headers={
                "Authorization": f"Bearer {od_token}",
                "Content-Type": "application/json",
            },
            json={"type": "view", "scope": "organization"},
            timeout=30,
        )
        if link_resp.status_code in (200, 201):
            web_url = link_resp.json().get("link", {}).get("webUrl", "")
            if web_url:
                log.info(f"  ↑ Uploaded {filename!r} → {web_url}")
                return web_url
    except Exception as exc:
        log.warning(f"createLink failed for {filename!r}: {exc}")

    fallback = item.get("webUrl", "")
    if fallback:
        log.warning(f"Using item webUrl for {filename!r}: {fallback}")
        return fallback

    return None


def upload_attachments(od_token: str, mail_token: str, msg: dict) -> str:
    if SKIP_UPLOADS:
        log.debug("SKIP_UPLOADS=true — skipping attachment upload.")
        return ""

    if not msg.get("hasAttachments"):
        return ""

    url = (
        f"https://graph.microsoft.com/v1.0/users/{TARGET_MAILBOX}"
        f"/messages/{msg['id']}/attachments?$select=id,name,contentType"
    )
    try:
        resp = requests.get(
            url,
            headers={"Authorization": f"Bearer {mail_token}"},
            timeout=60,
        )
        resp.raise_for_status()
        attachments = resp.json().get("value", [])
    except Exception as exc:
        log.warning(f"Could not fetch attachments for {msg['id']}: {exc}")
        return ""

    links: list[str] = []
    for att in attachments:
        filename = att.get("name", "")
        if not filename or not _is_resume(filename):
            log.debug(f"  ↷ Skipping non-resume attachment: {filename!r}")
            continue
        att_id = att.get("id", "")
        if not att_id:
            links.append(filename)
            continue
        try:
            content = _fetch_attachment_content(mail_token, msg["id"], att_id)
            link = _upload_to_onedrive(od_token, filename, content)
            links.append(link if link else filename)
        except Exception as exc:
            log.error(f"Failed to upload {filename!r}: {exc}")
            links.append(filename)

    return ", ".join(links)


# ── Subject filter ─────────────────────────────────────────────────────────────

# Subjects containing any of these words (case-insensitive) are silently
# dropped — they are delivery-status / bounce notifications, not profiles.
SKIP_SUBJECT_RE = re.compile(
    r"\b(undelivered|undeliverable|delivery\s+failed|delivery\s+status"
    r"|delivery\s+notification|mail\s+delivery|returned\s+mail"
    r"|non[- ]?deliverable|bounced?|failed\s+delivery"
    r"|delivered)\b",
    re.IGNORECASE,
)

def _is_skippable_subject(subject: str) -> bool:
    """Return True if the subject is a delivery-status notification to skip."""
    return bool(SKIP_SUBJECT_RE.search(subject))


SUBJECT_RE = re.compile(
    r"(?<!\S)(?:\[External\]\s*:\s*)?(?P<code>[A-Za-z]{2,4})\s*:",
    re.IGNORECASE,
)

_JR_PATTERNS = [
    re.compile(r"[-|]\s*(?:jr\s*)?(?P<jr>\d{4,})\s*$",   re.IGNORECASE),
    re.compile(r"^\s*(?:jr\s*)?(?P<jr>\d{4,})\s*[-|]",   re.IGNORECASE),
    re.compile(r"\(\s*jr\s*(?P<jr>\d{4,})\s*\)",          re.IGNORECASE),
    re.compile(r"\(\s*(?P<jr>\d{4,})\s*\)",               re.IGNORECASE),
    re.compile(r"\bjr\s*(?P<jr>\d{4,})\b",                re.IGNORECASE),
    re.compile(r"\b(?P<jr>\d{4,})\b",                     re.IGNORECASE),
]

def _extract_jr_and_skill(rest: str) -> tuple[Optional[str], str]:
    rest = re.sub(r"^[-|%\s]+|[-|%\s]+$", "", rest.strip()).strip()
    jr_no = None
    for pat in _JR_PATTERNS:
        m = pat.search(rest)
        if m:
            jr_no = m.group("jr")
            s, e = m.span()
            rest = re.sub(r"^[-|%\s]+|[-|%\s]+$", "", (rest[:s] + " " + rest[e:]).strip()).strip()
            break
    return jr_no, re.sub(r"\s{2,}", " ", rest).strip()

def parse_subject(subject: str) -> Optional[tuple[str, str, str, Optional[str]]]:
    subject = subject.strip()
    m = SUBJECT_RE.search(subject)
    if not m:
        return None
    code = m.group("code").upper()
    rest = subject[m.end():].strip()
    company_name = COMPANY_CODES.get(code, code)
    jr_no, skill = _extract_jr_and_skill(rest)
    return code, company_name, skill, jr_no


# ── HTML table parser ──────────────────────────────────────────────────────────
def _strip_html(text: str) -> str:
    """Remove all HTML tags."""
    return re.sub(r"<[^>]+>", "", text)

def _clean_cell(text: str) -> str:
    """
    Strip HTML tags, decode HTML entities (e.g. &nbsp; → space),
    then collapse whitespace.
    """
    # 1. Decode HTML entities first (&nbsp; &amp; &lt; etc.)
    text = html_lib.unescape(text)
    # 2. Remove any remaining HTML tags
    text = _strip_html(text)
    # 3. Replace non-breaking spaces (U+00A0) with regular spaces
    text = text.replace("\u00a0", " ")
    # 4. Collapse all whitespace to a single space and strip
    return re.sub(r"\s+", " ", text).strip()

# ── Value sanitisers ───────────────────────────────────────────────────────────
# contact_number: keep only digits (and leading +). Strip spaces, dashes, etc.
# email_id      : accept only values that look like a real email address.
#                 Non-email values (phone numbers, junk) are discarded → None.

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]{2,}$")
_PHONE_RE = re.compile(r"^[\d\s\-\+\(\)]{7,15}$")

def _looks_like_email(val: str) -> bool:
    return bool(val and _EMAIL_RE.match(val.strip()))

def _looks_like_phone(val: str) -> bool:
    """True if the value is all digits/spaces/dashes with no @ sign."""
    return bool(val and _PHONE_RE.match(val.strip()) and "@" not in val)

def _sanitize_contact(val: Optional[str]) -> Optional[str]:
    """
    Keep only the numeric digits (and a leading +) from a contact number.
    Strips spaces, dashes, dots, parentheses and any text.
    Returns None if nothing numeric remains.

    Examples:
      '9535389410 &nbsp;'  → '9535389410'
      '+91-98765-43210'    → '+9198765432​10'
      'chauhanjay@g.com'   → None  (email address — reject entirely)
      'N/A'                → None
    """
    if not val:
        return None
    v = val.strip()
    # Reject outright if it looks like an email — prevents email→contact bleed
    if "@" in v:
        log.warning(f"contact_number contains email address — discarding: {v!r}")
        return None
    # Strip everything except digits and a leading +
    digits = re.sub(r"[^\d+]", "", v)
    # Remove any + that's not at the very start
    digits = re.sub(r"(?<!^)\+", "", digits)
    if len(digits) < 7:
        # Too short to be a real phone number
        log.warning(f"contact_number has too few digits after cleaning — discarding: {v!r}")
        return None
    return digits or None

def _sanitize_email(val: Optional[str]) -> Optional[str]:
    """
    Accept only valid email addresses (must contain @ and a real domain).
    Rejects plain phone numbers, junk text, and partial values.
    Returns None if the value is not a recognisable email.

    Also strips any leading/trailing non-alphanumeric junk (e.g. residual
    &nbsp; that slipped through before HTML decoding) so that values like
    '&nbsp;agkakde7@gmail.com' are cleaned to 'agkakde7@gmail.com'.

    Examples:
      'john@gmail.com'          → 'john@gmail.com'
      '&nbsp;agkakde7@gmail.com'→ 'agkakde7@gmail.com'
      '9535389410'              → None  (phone number — reject)
      'john@'                   → None  (incomplete)
      'not-an-email'            → None
    """
    if not val:
        return None
    import html as _html
    # Decode any residual HTML entities, then strip non-printable/junk chars
    v = _html.unescape(val).strip()
    # Remove leading non-word characters before the local part (e.g. stray nbsp)
    v = re.sub(r"^[^\w]+", "", v).strip().lower()
    if _EMAIL_RE.match(v):
        return v
    log.warning(f"email_id value is not a valid email — discarding: {val!r}")
    return None

def _fix_swapped_contact_email(record: dict) -> dict:
    """
    Detect and correct the common case where email ends up in contact_number
    and a phone number ends up in email_id — caused by ambiguous column headers
    or column-order mismatches in the source HTML table.

    Cases handled:
      1. email in contact_number + phone in email_id  → swap both
      2. email in contact_number + email_id empty     → move to email_id, clear contact_number
      3. phone in email_id + contact_number empty     → move to contact_number, clear email_id
    """
    contact = (record.get("contact_number") or "").strip()
    email   = (record.get("email_id") or "").strip()

    contact_is_email = _looks_like_email(contact)
    contact_is_phone = _looks_like_phone(contact)
    email_is_phone   = _looks_like_phone(email)
    email_is_email   = _looks_like_email(email)

    # Case 1: both fields contain the wrong type — swap them
    if contact_is_email and email_is_phone:
        log.warning(
            f"Swapped contact/email detected — fixing: "
            f"contact_number={contact!r}, email_id={email!r}"
        )
        record["contact_number"] = email
        record["email_id"]       = contact
        return record

    # Case 2: email in contact_number, email_id is blank
    if contact_is_email and not email:
        log.warning(
            f"Email found in contact_number with blank email_id — moving: {contact!r}"
        )
        record["email_id"]       = contact
        record["contact_number"] = None
        return record

    # Case 3: phone in email_id, contact_number is blank
    if email_is_phone and not contact:
        log.warning(
            f"Phone found in email_id with blank contact_number — moving: {email!r}"
        )
        record["contact_number"] = email
        record["email_id"]       = None
        return record

    return record


def parse_html_table(html: str) -> list[dict]:
    rows: list[dict] = []
    for tbl in re.findall(r"<table[^>]*>(.*?)</table>", html, re.DOTALL | re.IGNORECASE):
        all_rows = re.findall(r"<tr[^>]*>(.*?)</tr>", tbl, re.DOTALL | re.IGNORECASE)
        if not all_rows:
            continue
        raw_headers = re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", all_rows[0], re.DOTALL | re.IGNORECASE)
        headers = [_clean_cell(h) for h in raw_headers]
        col_map = {i: _resolve_header(h) for i, h in enumerate(headers) if _resolve_header(h)}
        for row_html in all_rows[1:]:
            cells = re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", row_html, re.DOTALL | re.IGNORECASE)
            if not cells:
                continue
            cleaned = [_clean_cell(c) for c in cells]
            if all(v == "" for v in cleaned):
                continue
            record = {col_map[i]: v for i, v in enumerate(cleaned) if i in col_map}
            if record:
                rows.append(record)
    return rows


# ── Misc helpers ───────────────────────────────────────────────────────────────
def _extract_address(addr_obj: dict) -> str:
    try:
        return addr_obj["emailAddress"]["address"].strip().lower()
    except (KeyError, TypeError):
        return ""

def _normalize_name_from_email(email: str) -> str:
    if not email or "@" not in email:
        return ""
    local = re.sub(r"\d+", "", email.split("@", 1)[0])
    local = re.sub(r"[._\-]+", " ", local)
    return re.sub(r"\s+", " ", local).strip().title()

_DATE_FORMATS = ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y")

def _parse_date(raw: str) -> Optional[date]:
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw.strip(), fmt).date()
        except ValueError:
            continue
    return None

_REQUIRED_FIELDS = (
    "name_of_candidate", "contact_number", "email_id",
    "general_skill", "company_name", "recruiter", "email_from", "email_to",
)

def _record_status(data: dict) -> str:
    missing = [f for f in _REQUIRED_FIELDS if not data.get(f)]
    return "Fail" if missing else "Pass"

def _t(val) -> Optional[str]:
    if val is None:
        return None
    v = str(val).strip()
    return v or None


# ── Database helpers ───────────────────────────────────────────────────────────
def _make_insert_sql() -> pgsql.Composed:
    return pgsql.SQL(
        "INSERT INTO {schema}.{table} ({fields}) VALUES ({placeholders})"
    ).format(
        schema=pgsql.Identifier(DB_SCHEMA),
        table=pgsql.Identifier(DB_TABLE),
        fields=pgsql.SQL(", ").join(map(pgsql.Identifier, DB_COLUMNS)),
        placeholders=pgsql.SQL(", ").join(pgsql.Placeholder() * len(DB_COLUMNS)),
    )


def _verify_table_columns(conn) -> None:
    query = """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = %s AND table_name = %s
    """
    with conn.cursor() as cur:
        cur.execute(query, (DB_SCHEMA, DB_TABLE))
        existing = {row[0] for row in cur.fetchall()}

    if not existing:
        log.error(
            f"Table {DB_SCHEMA}.{DB_TABLE} not found or no columns returned. "
            f"Check DB_SCHEMA / DB_TABLE_NAME env vars and DB permissions."
        )
        raise SystemExit(1)

    missing = [c for c in DB_COLUMNS if c not in existing]
    if missing:
        log.error(
            f"Columns missing from {DB_SCHEMA}.{DB_TABLE}: {missing}\n"
            f"  Fix: add the missing columns to the table, or remove them from DB_COLUMNS."
        )
        raise SystemExit(1)

    log.info(
        f"Column pre-flight passed — all {len(DB_COLUMNS)} columns present "
        f"in {DB_SCHEMA}.{DB_TABLE}."
    )


def insert_records_incremental(conn, records: list[dict]) -> int:
    insert_sql = _make_insert_sql()
    inserted = 0
    for rec in records:
        row = tuple(rec.get(col) for col in DB_COLUMNS)
        try:
            with conn.cursor() as cur:
                cur.execute(insert_sql, row)
            conn.commit()
            inserted += 1
        except Exception as exc:
            log.error(
                f"Row insert failed for candidate "
                f"{rec.get('name_of_candidate', '?')!r}: {exc}"
            )
            conn.rollback()
    return inserted


# ── Excel writer ───────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

COL_WIDTHS = {
    "email_subject": 40, "recruiter": 18, "client_recruiter": 18,
    "email_from": 28, "email_to": 28, "delivery_type": 14,
    "company_name": 16, "general_skill": 22, "date": 12, "jr_no": 10,
    "name_of_candidate": 22, "contact_number": 16, "email_id": 28,
    "total_experience": 14, "relevant_experience": 16,
    "current_ctc": 14, "expected_ctc": 14, "notice_period": 14,
    "current_org": 22, "current_location": 18, "preferred_location": 18,
    "attachment": 35, "remarks": 30, "record_status": 14,
}

PRETTY_HEADERS = {
    "email_subject": "Email Subject", "recruiter": "Recruiter",
    "client_recruiter": "Client Recruiter", "email_from": "Email From",
    "email_to": "Email To", "delivery_type": "Delivery Type",
    "company_name": "Company", "general_skill": "General Skill",
    "date": "Date", "jr_no": "JR No",
    "name_of_candidate": "Candidate Name", "contact_number": "Contact Number",
    "email_id": "Email ID", "total_experience": "Total Exp",
    "relevant_experience": "Relevant Exp", "current_ctc": "Current CTC",
    "expected_ctc": "Expected CTC", "notice_period": "Notice Period",
    "current_org": "Current Org", "current_location": "Current Location",
    "preferred_location": "Preferred Location", "attachment": "Attachment",
    "remarks": "Remarks", "record_status": "Record Status",
}

def write_excel(records: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"
    ws.freeze_panes = "A2"

    for col_idx, col in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=PRETTY_HEADERS.get(col, col))
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = COL_WIDTHS.get(col, 16)

    ws.row_dimensions[1].height = 22

    STATUS_COLORS = {"Pass": "C6EFCE", "Fail": "FFC7CE"}

    for row_idx, rec in enumerate(records, 2):
        for col_idx, col in enumerate(EXCEL_COLUMNS, 1):
            val  = rec.get(col)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = CELL_FONT
            cell.alignment = LEFT
            if col == "record_status" and val in STATUS_COLORS:
                cell.fill = PatternFill("solid", start_color=STATUS_COLORS[val])
        ws.row_dimensions[row_idx].height = 16

    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    log.info(f"Saved {len(records)} row(s) → {output_path}")


# ── Main ───────────────────────────────────────────────────────────────────────
def run() -> None:
    log.info("=== Ad-hoc extractor starting ===")
    log.info(f"Target mailbox : {TARGET_MAILBOX}")
    log.info(f"OneDrive user  : {ONEDRIVE_USER}")
    log.info(f"OneDrive folder: {ONEDRIVE_ARCHIVE_FOLDER}")
    log.info(f"Pulling emails from: {SINCE_DATE} onwards")
    log.info(f"Folder: {INBOX_SUBFOLDER or '(root inbox)'}")
    log.info(f"DB target: {DB_SCHEMA}.{DB_TABLE}")
    log.info(f"Skip uploads: {SKIP_UPLOADS}")

    mail_token = get_mail_token()
    od_token   = get_onedrive_token()
    log.info("Tokens obtained.")

    if not SKIP_UPLOADS:
        _drive_check = requests.get(
            "https://graph.microsoft.com/v1.0/me/drive",
            headers={"Authorization": f"Bearer {od_token}"},
            timeout=30,
        )
        if _drive_check.status_code == 200:
            log.info(f"OneDrive access confirmed for {ONEDRIVE_USER}.")
            if ONEDRIVE_ARCHIVE_FOLDER:
                _ensure_onedrive_folder(od_token, ONEDRIVE_ARCHIVE_FOLDER)
        else:
            log.warning(
                f"OneDrive access check returned {_drive_check.status_code} — "
                f"attachment uploads may fail. Response: {_drive_check.text[:200]}"
            )
    else:
        log.info("Skipping OneDrive drive check (SKIP_UPLOADS=true).")

    try:
        db_conn = psycopg2.connect(DB_DSN)
        db_conn.autocommit = False
        log.info("DB connection opened.")
        _verify_table_columns(db_conn)
    except SystemExit:
        raise
    except Exception as exc:
        log.error(f"Cannot connect to DB: {exc} — records will NOT be inserted.")
        db_conn = None

    folder_id = resolve_folder_id(mail_token, INBOX_SUBFOLDER) if INBOX_SUBFOLDER else None
    emails    = fetch_bs_emails(mail_token, folder_id)

    records: list[dict] = []
    skipped  = 0
    db_count = 0

    for msg in emails:
        subject = msg.get("subject", "").strip()

        # Skip delivery-status / bounce notifications before anything else
        if _is_skippable_subject(subject):
            log.info(f"Skipping delivery-status email: {subject!r}")
            skipped += 1
            continue

        parsed  = parse_subject(subject)
        if parsed is None:
            skipped += 1
            continue

        company_code, company_name, general_skill, subject_jr_no = parsed

        from_addr = _extract_address(msg.get("from", {}))
        to_list   = msg.get("toRecipients", [])
        to_addr   = _extract_address(to_list[0]) if to_list else ""

        recruiter        = _normalize_name_from_email(from_addr)
        client_recruiter = _normalize_name_from_email(to_addr)
        delivery_type    = (
            "Internal"
            if VOLIBITS_DOMAIN in from_addr and VOLIBITS_DOMAIN in to_addr
            else "External"
        )

        attachment_str = upload_attachments(od_token, mail_token, msg)

        body_html = (msg.get("body") or {}).get("content", "")
        rows      = parse_html_table(body_html)
        if not rows:
            log.warning(f"No table in email: {subject!r} — adding skeleton row.")
            rows = [{}]

        email_date: Optional[date] = None
        raw_dt = msg.get("receivedDateTime", "")
        if raw_dt:
            try:
                email_date = datetime.fromisoformat(raw_dt.replace("Z", "+00:00")).date()
            except ValueError:
                pass

        email_records: list[dict] = []

        for row in rows:
            row_date = _parse_date(row["date"]) if row.get("date") else None
            if row_date is None:
                row_date = email_date or date.today()

            effective_jr_no = _t(row.get("jr_no")) or subject_jr_no
            candidate_name  = _t(row.get("name_of_candidate"))
            # Sanitise: contact_number → digits only; email_id → valid email only
            contact_number  = _sanitize_contact(_t(row.get("contact_number")))
            email_id_val    = _sanitize_email(_t(row.get("email_id")))

            record = {
                "email_subject":       subject,
                "recruiter":           recruiter,
                "client_recruiter":    client_recruiter,
                "email_from":          from_addr,
                "email_to":            to_addr,
                "delivery_type":       delivery_type,
                "company_name":        company_name,
                "general_skill":       _t(general_skill),
                "date":                row_date,
                "jr_no":               effective_jr_no,
                "name_of_candidate":   candidate_name,
                "contact_number":      contact_number,
                "email_id":            email_id_val,
                "total_experience":    _t(row.get("total_experience")),
                "relevant_experience": _t(row.get("relevant_experience")),
                "current_ctc":         _t(row.get("current_ctc")),
                "expected_ctc":        _t(row.get("expected_ctc")),
                "notice_period":       _t(row.get("notice_period")),
                "current_org":         _t(row.get("current_org")),
                "current_location":    _t(row.get("current_location")),
                "preferred_location":  _t(row.get("preferred_location")),
                "attachment":          _t(attachment_str),
                "remarks":             _t(row.get("remarks")),
            }

            # ── Fix swapped contact/email fields before status check ───────
            record = _fix_swapped_contact_email(record)

            record["record_status"] = _record_status({
                "name_of_candidate": record.get("name_of_candidate"),
                "contact_number":    record.get("contact_number"),
                "email_id":          record.get("email_id"),
                "general_skill":     _t(general_skill),
                "company_name":      company_name,
                "recruiter":         recruiter,
                "email_from":        from_addr,
                "email_to":          to_addr,
            })

            records.append(record)
            email_records.append(record)
            log.info(f"  + {candidate_name or '(unknown)'} | {subject!r} | {row_date}")

        if email_records and db_conn is not None:
            n = insert_records_incremental(db_conn, email_records)
            db_count += n
            log.info(
                f"  DB committed {n}/{len(email_records)} row(s) for: {subject!r} "
                f"(running total: {db_count})"
            )

    if db_conn is not None:
        try:
            db_conn.close()
            log.info("DB connection closed.")
        except Exception:
            pass

    log.info(f"Skipped {skipped} email(s) with no recognized company code.")
    log.info(f"Total rows collected  : {len(records)}")
    log.info(f"Total DB rows inserted: {db_count}")

    if not records:
        log.warning("No matching records found. Nothing to write.")
        return

    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"adhoc_candidates_{ts}.xlsx"
    write_excel(records, output_path)

    print(
        f"\n✅  Done — {len(records)} row(s) processed.\n"
        f"    DB rows inserted : {db_count}\n"
        f"    Excel output     : {output_path}\n"
    )


if __name__ == "__main__":
    run()